import openpyxl
import numpy as np

def gen_RES_test(RES_test, events, evMaxVal, output_path):

    num_events = len(events)

    wb = openpyxl.Workbook()    # 產生一個空的excel檔
    del wb['Sheet']

    # 取得測試資料及預測資料的列表
    all_tests = []
    all_predicts = []
    for i in range(len(RES_test)):
        all_tests.append(np.array(RES_test[i][0], dtype=np.float64).T.flatten().tolist()) # Observation
        all_predicts.append(np.array(RES_test[i][1], dtype=np.float64).T.flatten().tolist()) # Estimation

    # 將所有資料取到小數點第二位
    for i in range(len(all_tests)):
        for j in range(len(all_tests[i])):
            all_tests[i][j] = round(all_tests[i][j], 2)
            all_predicts[i][j] = round(all_predicts[i][j], 2)

    # 事件數據工作表(第一個工作表)
    sheet = wb.create_sheet("Event Statistics")
    sheet.append([ "Number of event (Nev)" ])
    sheet.append([ num_events ])
    sheet.append([ "Event with max value" ])
    sheet.append([ evMaxVal ]) 
    sheet.append([ "Event", "number of data" ])
    for i in range(num_events):
        sheet.append([i+1, len(events[i])])

    # 將每個Event的預測結果分成不同工作表(第二個~倒數第二個工作表)
    for i in range(num_events):
        sheet_name = "Event " + "%02d" %(i+1)
        sheet = wb.create_sheet(sheet_name)
        sheet.freeze_panes = "A2"
        sheet.append([ 'Observation', 'Estimation', 'Event' ])
        for j in range(len(all_tests[i])):
            sheet.append([ all_tests[i][j], all_predicts[i][j] , i+1 ])

    # 將所有預測結果存到同一個工作表(最後一個工作表)
    sheet = wb.create_sheet("All Events")
    sheet.freeze_panes = "A2"
    sheet.append([ 'Observation', 'Estimation', 'Event' ])
    for i in range(num_events):
        for j in range(len(all_tests[i])):
            sheet.append([ all_tests[i][j], all_predicts[i][j], i+1 ])

    # 輸出excel檔
    wb.save(output_path)
    print("Result saved in " + output_path)
